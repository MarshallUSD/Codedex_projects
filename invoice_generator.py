# FinTrack Co. --- Multi Format Invoice Generator

from abc import ABC, abstractmethod
from typing import List
from datetime import datetime
import os


# ============================================================================
# 1️⃣ ABSTRACT BASE CLASS
# ============================================================================

class InvoiceGenerator(ABC):
    """
    Abstract base class for invoice generation.
    Defines the interface that all concrete generators must implement.
    """
    
    def __init__(self, client_name: str, items: List[dict]):
        """
        Initialize invoice generator.
        
        Args:
            client_name: Name of the client
            items: List of dictionaries with 'name' and 'price' keys
        """
        self.client_name = client_name
        self.items = items
        self.generation_date = datetime.now()  # Fixed: was 'generator_date'
    
    def calculate_total(self) -> float:
        """
        Calculate the total sum of all items.
        
        Returns:
            float: Total amount
        """
        return sum(item['price'] for item in self.items)
    
    @abstractmethod
    def generate_invoice(self) -> str:
        """
        Abstract method to generate invoice in specific format.
        Must be overridden by all subclasses.
        
        Returns:
            str: Path to the generated invoice file
        """
        pass


# ============================================================================
# 2️⃣ CONCRETE IMPLEMENTATIONS
# ============================================================================

class PDFInvoiceGenerator(InvoiceGenerator):
    """Generate invoices in PDF format using ReportLab"""
    
    def generate_invoice(self) -> str:
        """
        Generate a PDF invoice file.
        
        Returns:
            str: Path to generated PDF file
        """
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        
        # Create output directory
        os.makedirs('invoices', exist_ok=True)
        
        # Fixed: %D should be %d for day of month
        filename = f"invoices/invoice_{self.client_name.replace(' ', '_')}_{self.generation_date.strftime('%Y%m%d_%H%M%S')}.pdf"
        
        # Fixed: was 'docs', should be 'doc'
        doc = SimpleDocTemplate(filename, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Invoice Title
        title = Paragraph("<b>INVOICE</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        # Fixed: was 'style', should be 'styles'
        client_info = Paragraph(f"<b>Client:</b> {self.client_name}", styles['Heading2'])
        elements.append(client_info)
        
        date_info = Paragraph(
            f"<b>Generated On:</b> {self.generation_date.strftime('%Y-%m-%d %H:%M:%S')}", 
            styles['Normal']
        )
        elements.append(date_info)
        elements.append(Spacer(1, 0.3*inch))
        
        # Items table
        table_data = [['Item Name', 'Price']]
        for item in self.items:
            table_data.append([item['name'], f"${item['price']:.2f}"])
        
        # Add total row
        table_data.append(['TOTAL', f"${self.calculate_total():.2f}"])
        
        # Create and style table
        items_table = Table(table_data, colWidths=[4*inch, 2*inch])
        items_table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Body styling
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('GRID', (0, 0), (-1, -2), 1, colors.black),
            
            # Total row styling
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 12),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
        ]))
        
        elements.append(items_table)
        
        # Build PDF
        doc.build(elements)
        
        return filename


class ExcelInvoiceGenerator(InvoiceGenerator):
    """Generate invoices in Excel format using openpyxl"""
    
    def generate_invoice(self) -> str:
        """
        Generate an Excel (.xlsx) invoice file.
        
        Returns:
            str: Path to generated Excel file
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Create output directory
        os.makedirs('invoices', exist_ok=True)
        
        # Generate filename
        filename = f"invoices/invoice_{self.client_name.replace(' ', '_')}_{self.generation_date.strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"
        
        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        
        # Title
        ws['A1'] = "INVOICE"
        ws['A1'].font = Font(size=18, bold=True)
        ws.merge_cells('A1:B1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Client information
        ws['A3'] = "Client:"
        ws['A3'].font = Font(bold=True)
        ws['B3'] = self.client_name
        
        # Generation date
        ws['A4'] = "Generated On:"
        ws['A4'].font = Font(bold=True)
        ws['B4'] = self.generation_date.strftime('%Y-%m-%d %H:%M:%S')
        
        # Column headers
        ws['A6'] = "Item Name"
        ws['B6'] = "Price"
        ws['A6'].font = Font(bold=True, color='FFFFFF')
        ws['B6'].font = Font(bold=True, color='FFFFFF')
        ws['A6'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws['B6'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws['A6'].alignment = Alignment(horizontal='center')
        ws['B6'].alignment = Alignment(horizontal='center')
        
        # Add items
        row = 7
        for item in self.items:
            ws[f'A{row}'] = item['name']
            ws[f'B{row}'] = item['price']
            ws[f'B{row}'].number_format = '$#,##0.00'
            row += 1
        
        # Add total row
        total_row = row
        ws[f'A{total_row}'] = "TOTAL"
        ws[f'A{total_row}'].font = Font(bold=True, size=12)
        ws[f'B{total_row}'] = self.calculate_total()
        ws[f'B{total_row}'].number_format = '$#,##0.00'
        ws[f'B{total_row}'].font = Font(bold=True, size=12)
        ws[f'B{total_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_num in range(6, total_row + 1):
            ws[f'A{row_num}'].border = thin_border
            ws[f'B{row_num}'].border = thin_border
        
        # Align prices to right
        for row_num in range(7, total_row + 1):
            ws[f'B{row_num}'].alignment = Alignment(horizontal='right')
        
        # Save workbook
        wb.save(filename)
        
        return filename


class HTMLInvoiceGenerator(InvoiceGenerator):
    """Generate invoices in HTML format"""
    
    def generate_invoice(self) -> str:
        """
        Generate an HTML invoice file.
        
        Returns:
            str: Path to generated HTML file
        """
        # Create output directory
        os.makedirs('invoices', exist_ok=True)
        
        # Generate filename
        filename = f"invoices/invoice_{self.client_name.replace(' ', '_')}_{self.generation_date.strftime('%Y%m%d_%H%M%S')}.html"
        
        # Build HTML content
        html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Invoice - {self.client_name}</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            max-width: 800px;
            margin: 40px auto;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .invoice {{
            background-color: white;
            padding: 40px;
            box-shadow: 0 0 10px rgba(0,0,0,0.1);
        }}
        h1 {{
            text-align: center;
            color: #333;
            border-bottom: 3px solid #333;
            padding-bottom: 20px;
        }}
        .info {{
            margin: 30px 0;
        }}
        .info p {{
            margin: 10px 0;
            font-size: 14px;
        }}
        .info strong {{
            display: inline-block;
            width: 120px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 30px 0;
        }}
        th {{
            background-color: #366092;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
        }}
        td {{
            padding: 12px;
            border-bottom: 1px solid #ddd;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        .price {{
            text-align: right;
        }}
        .total-row {{
            background-color: #ffff99 !important;
            font-weight: bold;
            font-size: 1.2em;
            border-top: 2px solid #333;
        }}
    </style>
</head>
<body>
    <div class="invoice">
        <h1>INVOICE</h1>
        
        <div class="info">
            <p><strong>Client:</strong> {self.client_name}</p>
            <p><strong>Generated On:</strong> {self.generation_date.strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Item Name</th>
                    <th class="price">Price</th>
                </tr>
            </thead>
            <tbody>
"""
        
        # Add items
        for item in self.items:
            html_content += f"""                <tr>
                    <td>{item['name']}</td>
                    <td class="price">${item['price']:.2f}</td>
                </tr>
"""
        
        # Add total
        html_content += f"""                <tr class="total-row">
                    <td>TOTAL</td>
                    <td class="price">${self.calculate_total():.2f}</td>
                </tr>
            </tbody>
        </table>
    </div>
</body>
</html>
"""
        
        # Write to file
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return filename


# ============================================================================
# 3️⃣ BUSINESS LOGIC LAYER
# ============================================================================

class InvoiceManager:
    
    
    def __init__(self, generator: InvoiceGenerator):
        
        self.generator = generator
    
    def export_invoice(self) -> None:
    
        print(f"Generating invoice for client: {self.generator.client_name}")
        print(f"Total amount: ${self.generator.calculate_total():.2f}")
        
        filepath = self.generator.generate_invoice()
        
        print(f"✓ Invoice successfully generated: {filepath}")
        print()

def main():
    """
    Main execution scenario demonstrating invoice generation in multiple formats.
    """
    print("=" * 70)
    print("FINTRACK CO. - INVOICE GENERATION SYSTEM")
    print("=" * 70)
    print()
    
    
    client_name = "Acme Corporation"
    items = [
        {'name': 'Web Development Services', 'price': 5000.00},
        {'name': 'UI/UX Design', 'price': 2500.00},
        {'name': 'If it changed, this is marshall', 'price': 7500.00},
        {'name': 'Cloud Hosting (Annual)', 'price': 1200.00},
        {'name': 'Technical Support', 'price': 800.00}
    ]
    
    print(f"Client: {client_name}")
    print(f"Number of items: {len(items)}")
    print()
    print("-" * 70)
    print()
    
    # 1. Generate PDF Invoice
    print("📄 GENERATING PDF INVOICE")
    print("-" * 70)
    pdf_generator = PDFInvoiceGenerator(client_name, items)
    pdf_manager = InvoiceManager(pdf_generator)
    pdf_manager.export_invoice()
    
    # 2. Generate Excel Invoice
    print("📊 GENERATING EXCEL INVOICE")
    print("-" * 70)
    excel_generator = ExcelInvoiceGenerator(client_name, items)
    excel_manager = InvoiceManager(excel_generator)
    excel_manager.export_invoice()
    
    # 3. Generate HTML Invoice
    print("🌐 GENERATING HTML INVOICE")
    print("-" * 70)
    html_generator = HTMLInvoiceGenerator(client_name, items)
    html_manager = InvoiceManager(html_generator)
    html_manager.export_invoice()
    
    print("=" * 70)
    print("ALL INVOICES GENERATED SUCCESSFULLY!")
    print("=" * 70)
    print()
    print("Files saved in the 'invoices' directory")
    print()
    print("💡 Key OOP Concepts Demonstrated:")
    print("   • Abstract Base Class (ABC) - InvoiceGenerator")
    print("   • Polymorphism - InvoiceManager works with any generator")
    print("   • Inheritance - PDF, Excel, HTML generators extend base class")
    print("   • Encapsulation - Business logic separated from format logic")
    print("   • Open/Closed Principle - Easy to add new formats without modifying existing code")


if __name__ == "__main__":
    main()